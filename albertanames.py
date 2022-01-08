'''*****************************************************************************
FILE: albertaNames.py
Author: asad70
-------------------------------------------------------------------
****************************************************************************'''

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import pprint 
import pickle


def open_file(fName):
    '''
    Expects the filename string (including ".xlsx" suffix) of a Excel .xlsx 
    file, structured as follows:
    1	    Michael	   732	           Boy	        1980
    2	    Christopher	   633	           Boy	        1980 <--Start At line 7
    :       :              :               :             :
    2,175   Eloise         1               Girl         2018
    
    extracts the data, starting at row 7, find the latest year in the data. 
    Then return the raw data in the form of a list of lists where each list is
    a row of data from the Excel, i.e. [rank, name, frequency, gender, year]. 
    The latest year in the data must be an integer. The first 6 rows must not 
    be included.The raw data list of lists and the latest year are
    returned.
    In case of failure to open the file this returns NONE.
     Parameters:
        fName: Name of the file
    Return:
        data: list 
              data is in the form of a list of lists where each list is
    a row of data from the Excel, i.e. [rank, name, frequency, gender, year]. 
        max_year: int
                 latest year in the data
    '''
    
    try:
        wk_bk = load_workbook(fName, data_only = True)
    except:
        print("File does not exist")
        return
      
    
    sheet = wk_bk.active
    row_max, col_max = sheet.max_row, sheet.max_column  
    
    col_max_letter = chr(ord('A') + sheet.max_column - 1)   # Last column: 'E'
    col_last = chr(ord('A') + sheet.max_column - 1)  
    cell_last = col_last + str(row_max)
    max_cell = col_max_letter + str(sheet.max_row)
    
    data = [ ]
    rows = sheet['A3': cell_last]       # To be returned, skips first 2 lines
    
    # Go thru the rows, put all of the row data in alist and append that
    # list to 
    for row in rows:                    # Iterate vertically over rows
        row_data = [ ]                  # New list for row
        for cell in row:                # Iterate over every element in a row
            if isinstance(cell.value, str) and cell.value[0] == '=':
                row_data.append(data[len(data)-1][0])                
            else:            
                row_data.append(cell.value)
        row_data[-1] = str(row_data[-1])
        data.append(row_data)           # Add it to final list
        		
    max_year = data[-1][-1]
     
    return data, int(max_year)
    
    
def create_names_dict(data, names):    
    '''
    Expects the raw data extracted from the Excel file in open_file and an 
    empty dictionary as parameters. This function processes the raw data and 
    returns the dictionary using names as keys and a list of lists each 
    containing frequency, gender, and year as the value. This dictionary is 
    returned. 
    For example:
    {..., 'Kelly': [[69,'Boy',1980], 
                   [61, 'Boy', 1981], 
                   [53, 'Boy', 1982],
                   [37, 'Boy', 1983], 
                   [32, 'Boy', 1984], 
                   [123, 'Girl', 1980], ...}
    Parameters:
        data: raw data
        names: empty dictionary
    
    Return:
        names: dictionary 
               containing names as keys and a list of lists  each 
        containing frequency, gender, and year as the value.
    '''
    
    # Coverts year to int from str.
    count = 0
    for i in data:
        data[count][-1] = int(data[count][-1])
        count += 1
    
    for i in range(len(data)):
        name = data[i][1]        
        # If name isn't in names, add that to dict as key value
        if name not in names.keys():
            names[name] = []          
            names[name] = [data[i][2:]] # [data[i][2:]] = frequency, gender, yr
        else:
            names[name].append(data[i][2:]) # [32, 'Boy', 1984]
    return names

def create_top_ten_dict(data, names):
    '''
    Expects raw data extracted from the Excel file in open_file and an empty 
    dictionary as parameters. This function processes the raw data and returns 
    the dictionary using years as keys and a list of lists each containing 
    rank, name, gender, and year as the value. For example:
    {1980: [[1, 'Michael', 732, 'Boy'], [2, 'Christopher', 633, 'Boy'], 
    [3, 'David', 537, 'Boy']...}
    
    Parameters:
        data: raw data
        names: empty dictionary
    
    Return:
        names: dictionary 
               containing years as keys and a list of lists each 
        containing rank, name, gender, and year as the value.
    
        '''
    
    for i in range(len(data)):
        rank = data[i][0]
        year = int(data[i][4])    
        
        if isinstance(rank, int) == True and rank < 11:
            if year in names.keys():
                names[year].append(data[i][:4]) #[1, 'Michael', 732, 'Boy']
            else:
                names[year] = []            
                names[year] = [data[i][:4]] # rank, name, gender, and year                  
    return names

def load_file(fName, names, top_ten):
    '''
    Expects a default filename, the names dictionary, and the top ten 
    dictionary as parameters. If the user enters a filename, this function  
    attempts to open the file. If no filename is entered, this function 
    attempts to open the file with the default filename. If the user presses 
    enter without entering a file name, it uses the default (which must be
    baby_names_frequency.xlsx).
    In case of failure to open the specified file this returns NONE.
    
    Parameters:
        fileName: Name of the default file
        names: names dictionary
        top_ten: dictionary of top ten lists
    
    Return: create_names_dict(data, names): dictionary 
               containing names as keys and a list of lists  each containing 
               frequency, gender, and year as the value.
        create_top_ten_dict(data, name): dictionary containing years as 
            keys and a list of lists each containing rank, name, gender, and 
            year as the value.
        max_year: int 
               latest year
    '''    
    

    wk_bk = input("Enter a file name [baby-names-frequency-80-84.xlsx]: ")
    if len(wk_bk) > 4 and ".xlsx" in wk_bk.strip()[-5:]:
        wk_bk = wk_bk
    if len(wk_bk) == 0:
        wk_bk = 'baby-names-frequency-80-84.xlsx'
        
    try:
        data, max_year = open_file(wk_bk) 
        print("Data has been loaded and processed")
    except: 
        return 
    
    name = {} # initializing to pass empty dic to create_top_ten below.
    return (create_names_dict(data, names), create_top_ten_dict(data, name), 
            max_year)
     
def pickle_helper(data):
    '''
    Expects data. If the user nters a filename, the dictionaries are to be 
    saved with that filename. If no filename is entered, the default filename 
    (baby_names.p) is used to save the dictionaries. If the user presses enter 
    without entering a file name, it uses the default (which is be 
    baby_names.p).
    This function calls pickle_dicts(fName, names, top_ten, max_year) to save 
    the data.
    
    Parameters:
        data: tuple
              containing names, top_ten and max_year
    
    Return: NONE
    '''    
    names, top_ten, max_year = data    
    
    if len(names) == 0:
        if len(top_ten) > 0: pass             # pass is a null operator.
        else:
            print("There are no data")
            return  
        
    wk_bk = input("Enter a file name [baby_names.p]: ")    
    
    if len(wk_bk) > 0 and wk_bk.strip()[-2:] == ".p":
        wk_bk = wk_bk    
    else:
        wk_bk = "baby_names.p"    
    pickle_dicts(wk_bk, names, top_ten, max_year)


def pickle_dicts(fName, names, top_ten, max_year):
    '''
    Expects these parameters: a default filename, the dictionary of names, the
    dictionary of top ten lists, and the latest year as parameters. If the user 
    enters a filename, the dictionaries are to be saved with that filename. If 
    no filename is entered, the default filename (baby_names.p) is used to save 
    the dictionaries. If the user presses enter without entering a file name, 
    it uses the default (which is be baby_names.p).
    
    Parameters:
        fName: Name of the default file
        names: names dictionary
        top_ten:  dictionary of top ten lists
        max_year: latest year 
    
    Return: NONE
    '''
    if len(names) == 0:
        if len(top_ten) > 0: pass
        else:
            print("There are no data")
            return       
    try:
        wb_loaded = open(fName, "wb")
    except:
        wb_loaded = open("baby_names.p", "wb")
        
    data = (names, top_ten, max_year)
    pickle.dump(data, wb_loaded)
    wb_loaded.close() 
    print("Saved pickled data in " + fName + ".")
    
def load_helper():
    '''
    This function asks user for file that us used by load_pickle to load file.
    If the user enters a filename, this function calls load_pickle(fName) to 
    open the file otherwise the default filename is used. If the user presses 
    enter without entering a file name, it uses the default (which is 
    baby_names.p). This function returns a tuple containing the dictionaries 
    for the names, the top ten lists, and the latest year.

    Parameters:
        NONE
    Return: 
    Tuples that contain the following:
        names: dictionary
            dictionaries for the name
        top: list
            top ten lists
        year: int
              latest year
    '''    
    
    wk_bk = input("Enter a file name [baby_names.p]: ")     
    if len(wk_bk) == 0:
        wk_bk = "baby_names.p"
    elif len(wk_bk) > 0 and wk_bk.strip()[-2:] == ".p":
        wk_bk = wk_bk    
    
    try:
        names, top_ten, max_year = load_pickle(wk_bk)
    except: return
    
    return names, top_ten, max_year

def load_pickle(fName):
    '''
    Expects a default filename as a parameter. If the user enters a filename, 
    this function attempt to open the file otherwise the default filename is 
    used. If the user presses enter without entering a file name, it uses the 
    default (which is baby_names.p). This function returns a tuple containing 
    the dictionaries for the names, the top ten lists, and the latest year.
    Parameters:
        fName: Name of the default file
    
    Return: 
    Tuples that contain the following:
        names: dictionary
            dictionaries for the name
        top: list
            top ten lists
        year: int
              latest year

    '''
        
    try:
        wb_loaded = open(fName, "rb")    
    except:
        print("Could not load pickle from " + fName + ".")  
        return 
        
    data = pickle.load(wb_loaded)    
    names, top_ten, max_year = data
    
    print("Loaded pickled data from " + fName + ".")
    return (names, top_ten, max_year)
    

def search_helper(full_list, search):
    '''
    Expects the dictionary of all names in which data sorted is with separate 
    years, each year containning the gender and frequency and the name to be 
    searched as parameters. This function prints out the frequencies of boys a
    and girls who were given that name in each year. If there were no babies 
    given the searched for name, a message is displayed that no babies were 
    given this name (capitalized).
    
    Parameters:
        full_list: dictionary
               dictionary of all names, data sorted with separate years.
        search: str
              name to be searched
    
    Return: NONE
    '''    
    # Dictionary sorted to prvent error while adding years in previous func.
    yr_dict = full_list[search]
    yr_dict = dict(sorted(yr_dict.items()))
    
    print("\n" + search + ":")
    print("\tBoys\tGirls")
    # for name in the full_list, that has the following data (as ex.)
    # {'Michael': {1980: [['Boy', 732], ['Girl', 705]]
    for list_name in full_list:
        if list_name == search:     # If the name is same as being searched
            # Acessing years in the keys.
            for years in (yr_dict.keys()):  
                # Try/except used to prevent 'index out of range error'
                try:
                    # Case where name is given to only boys.
                    # If Boy in {'Michael':{1980: [['Boy', 732], ['Girl', 705]]
                    if  'Boy' in yr_dict[years][0]: 
                        # if length of {1980: [['Boy', 732], ['Girl', 705]]} =1
                        if len(yr_dict[years]) == 1:
                            print(years, "\t", end = "")
                            print(yr_dict[years][0][1], end = "")                                  
                            print("\t", "0")
                                    
                    # Case where name is only given to girl
                    # Ex: full_list[search][years][0] = ['Boy', 732]
                    if 'Girl' in yr_dict[years][0]:
                        if len(yr_dict[years]) == 1:
                            print(years, "\t", "0", end = "")
                            print("\t", yr_dict[years][0][1])    
                        
                    # Case where name is given to both boys and girls
                    if 'Boy' in yr_dict[years][0] and \
                    'Girl' in yr_dict[years][1]:
                        print(years, "\t", end = "")
                        print(yr_dict[years][0][1], end = "")
                        print("\t" , yr_dict[years][1][1])                     
                except: pass    
                
def name_search(names, year):
    '''
    Expects the dictionary of all names, the latest year as parameters. This
    function asks the user for a name and calls name_trend(names, year) and 
    search_helper(full_list, search) to print out the frequencies of boys and girls 
    who were given that name in each year. If there were no babies given the 
    searched for name, a message is displayed that no babies were given this 
    name (capitalized).
    
    Parameters:
        names: dictionary
               dictionary of all names
        year: int
              latest year
    
    Return: NONE
    '''
    if len(names.keys()) == 0:
        print("There are no data")
        return
    
    search = input("Enter a name: ")
    search = search.capitalize()
    if search not in names.keys():
        print("There were no babies named " + search + " born in Alberta" 
              " between 1980 and " + str(year))
        return
            
    full_list = name_trend(names, year)
    
    # Passing the data to search helper for rest of the work.
    search_helper(full_list, search)   
    
    
def name_trend(names, year):
    '''
    Expects the dictionary of all names, the latest year as parameters. This
    function takes the dict and turn it into a dictionary of dictionary that has
    name as key with dict -year as key that has list of list of boys/girls freq.
    # Ex: {'Michael': {1980: [['Boy', 732], ['Girl', 705]]
    
    Parameters:
        names: dictionary
               dictionary of all names
        year: int
              latest year
    
    Return: full_list: dictionary of dict, with names as key and dict with year
                as key and boys/girls names with frequency as lists of lists.
                # Ex: {'Michael': {1980: [['Boy', 732], ['Girl', 705]]
    '''    
    full_list = {}                  # New dictionary
    for keys in names:              # keys refer to name in dictionary.
        full_list[keys] = {}        # add [key]-name dictionary to full-list 
        for name in names[keys]:     # name is the list data for single name.
            year = name[-1]
            # if year has not been added, then add.
            if year not in full_list[keys].keys():
                full_list[keys][year] = []
            else: pass
            gender = name[1]
            freq = name[0]
            # Ex: {'Michael': {1980: [['Boy', 732], ['Girl', 705]]
            full_list[keys][year].append([gender, freq])
    
    return full_list 


def open_file(fName):
    '''
    Expects the filename string (including ".xlsx" suffix) of a Excel .xlsx 
    file, structured as follows:
    1	    Michael	   732	           Boy	        1980
    2	    Christopher	   633	           Boy	        1980 <--Start At line 7
    :       :              :               :             :
    2,175   Eloise         1               Girl         2018
    
    extracts the data, starting at row 7, find the latest year in the data. 
    Then return the raw data in the form of a list of lists where each list is
    a row of data from the Excel, i.e. [rank, name, frequency, gender, year]. 
    The latest year in the data must be an integer. The first 6 rows must not 
    be included.The raw data list of lists and the latest year are
    returned.
    
     Parameters:
        fName: Name of the file
    Return:
        data: list 
              data is in the form of a list of lists where each list is
    a row of data from the Excel, i.e. [rank, name, frequency, gender, year]. 
        max_year: int
                 latest year in the data
    '''
    
    try:
        wk_bk = load_workbook(fName, data_only = True)
    except:
        print("File does not exist")
        return
      
    
    sheet = wk_bk.active
    row_max, col_max = sheet.max_row, sheet.max_column  
    
    col_max_letter = chr(ord('A') + sheet.max_column - 1)   # Last column: 'E'
    col_last = chr(ord('A') + sheet.max_column - 1)  
    cell_last = col_last + str(row_max)
    max_cell = col_max_letter + str(sheet.max_row)
    
    data = [ ]
    rows = sheet['A7': cell_last]       # To be returned, skips first 6 lines
    
    # Go thru the rows, put all of the row data in alist and append that
    # list to 
    for row in rows:                    # Iterate vertically over rows
        row_data = [ ]                  # New list for row
        for cell in row:                # Iterate over every element in a row
            if isinstance(cell.value, str) and cell.value[0] == '=':
                row_data.append(data[len(data)-1][0])                
            else:            
                row_data.append(cell.value)
        row_data[-1] = str(row_data[-1])
        data.append(row_data)           # Add it to final list
        		
    max_year = data[-1][-1]
     
    return data, int(max_year)
    
    
def create_names_dict(data, names):    
    '''
    Expects the raw data extracted from the Excel file in open_file and an 
    empty dictionary as parameters. This function processes the raw data and 
    returns the dictionary using names as keys and a list of lists each 
    containing frequency, gender, and year as the value. This dictionary is 
    returned. 
    For example:
    {..., 'Kelly': [[69,'Boy',1980], 
                   [61, 'Boy', 1981], 
                   [53, 'Boy', 1982],
                   [37, 'Boy', 1983], 
                   [32, 'Boy', 1984], 
                   [123, 'Girl', 1980], ...}
    Parameters:
        data: raw data
        names: empty dictionary
    
    Return:
        names: dictionary 
               containing names as keys and a list of lists  each 
        containing frequency, gender, and year as the value.
    '''
    
    # Coverts year to int from str.
    count = 0
    for i in data:
        data[count][-1] = int(data[count][-1])
        count += 1
    
    for i in range(len(data)):
        name = data[i][1]        
        # If name isn't in names, add that to dict as key value
        if name not in names.keys():
            names[name] = []          
            names[name] = [data[i][2:]] # [data[i][2:]] = frequency, gender, yr
        else:
            names[name].append(data[i][2:]) # [32, 'Boy', 1984]
    return names

def create_top_ten_dict(data, names):
    '''
    Expects raw data extracted from the Excel file in open_file and an empty 
    dictionary as parameters. This function processes the raw data and returns 
    the dictionary using years as keys and a list of lists each containing 
    rank, name, gender, and year as the value. For example:
    {1980: [[1, 'Michael', 732, 'Boy'], [2, 'Christopher', 633, 'Boy'], 
    [3, 'David', 537, 'Boy']...}
    
    Parameters:
        data: raw data
        names: empty dictionary
    
    Return:
        names: dictionary 
               containing years as keys and a list of lists each 
        containing rank, name, gender, and year as the value.
    
        '''
    
    for i in range(len(data)):
        rank = data[i][0]
        year = int(data[i][4])    
        
        if isinstance(rank, int) == True and rank < 11:
            if year in names.keys():
                names[year].append(data[i][:4]) #[1, 'Michael', 732, 'Boy']
            else:
                names[year] = []            
                names[year] = [data[i][:4]] # rank, name, gender, and year                  
    return names

def print_girl_names(g_data):
    '''
    This helper function of print_top_ten expects the list of list of top ten 
    names. This function prints out the top ten girls names in the 
    user-specified years.
    
    Parameters: b_data: list of list
                        lists of list of all the girls name in given year.
          Ex: [[1, 'Jennifer', 792, 'Girl'], [2, 'Amanda', 486, 'Girl']]

                          
    Return: NONE
    '''
       
    rank = {}                   # New empty dict
    for i in g_data:            # i - is single list in boys data
        if i[0] in rank.keys(): # i[0] is year - if year already in rank,append.
            rank[i[0]].append([i[1], i[2]])
        else:                   
            rank[i[0]] = {}     # Else: add year to dict. Ex {1:}
            rank[i[0]] = []     # Add list to that year. Ex: {1: []}
            rank[i[0]].append([i[1], i[2]]) # Ex: {1: [['Jennifer', 792]]}
       
    for num in rank:            # num is key in rank from (1-10).
        name = str(rank[num][0][0])
        freq = str(rank[num][0][1])   
        
        if (len(rank[num])) == 1:  # If ranking has only one name.
            print(str(num) + "\t" + name + ": " + freq + "\t")
        else:
            print(num, end = "")   # Else ranking has more than one name then:
            for i in rank[num]:    # print each separately
                print("\t" + str(i[0]) + ": " + str(i[1]), end = "")
                length = len(rank[num])
            print("\t")
            for i in range(length - 1):    # Printing the number on empty line.
                if (num + 1 + i) < 11:     # Ex: 6    Lisa: 264   Nicole: 264
                    print(str(num + 1 + i) + "\t")     #     7
                   
        
def print_boy_names(b_data):
    '''
    This helper function of print_top_ten expects the list of list of top ten 
    names. This function prints out the top ten boys names in the 
    user-specified years.
    
    Parameters: b_data: list of list
                        lists of list of all the boys name in given year.
          Ex: [[1, 'Michael', 792, 'Boy'], [2, 'Christopher', 486, 'Boy']]

                          
    Return: NONE
    '''    
                      
    rank = {}                   # New empty dict
    for i in b_data:            # i - is single list in boys data
        if i[0] in rank.keys(): # i[0] is year - if year already in rank,append.
            rank[i[0]].append([i[1], i[2]])
        else:                   
            rank[i[0]] = {}     # Else: add year to dict. Ex {1:}
            rank[i[0]] = []     # Add list to that year. Ex: {1: []}
            rank[i[0]].append([i[1], i[2]]) # Ex: {1: [['Jennifer', 792]]}
       
    for num in rank:
        name = str(rank[num][0][0])
        freq = str(rank[num][0][1])   
        if (len(rank[num])) == 1:  # If ranking has only one name.
            print(str(num) + "\t" + name + ": " + freq + "\t")
        else:
            print(num, end = "")   # Else ranking has more than one name then:
            for i in rank[num]:    # print each separately
                print("\t" + str(i[0]) + ": " + str(i[1]), end = "")
                length = len(rank[num])
            print("\t")
            for i in range(length - 1):    # Printing the number on empty line.
                if (num + 1 + i) < 11:     # Ex: 6    Lisa: 264   Nicole: 264
                    print(str(num + 1 + i) + "\t")     #     7
                   
            
def print_top_ten(names, max_year):
    '''
    Expects the dictionary of top ten lists and the latest year as parameters. 
    This function calls ask the user for a year, error checks to ensure that 
    the year is in the range from 1980 to the latest year, and calls 
    print_girl_names(names, year) and print_boy_names(names, year)
    prints the top ten list of names with their frequencies for that year. 
    
    Parameters:  names: dictionary
                          dictionary of top ten lists
                 max_year: int
                          latest year
                          
    Return: NONE 
    '''
    if len(names.keys()) == 0:
        print("There are no data")
        return    
    
    year = input("Enter year (1980 to " + str(max_year) + "): ")
    while True:
        if len(year) != 0 and year.isdigit():
            year = int(year)
            if 1980 <= year <= max_year:
                break
        year = input("Enter year (1980 to " + str(max_year) + "): ")
        
    b_data, g_data = [], []
    for data in names[year]: 
        if data[-1] == 'Boy':       
            b_data.append(data)     # b_data: have only the boys data    
        elif data[-1] == 'Girl':
            g_data.append(data)     # g_data: have only the girls data
           
    print("\nTop 10 names for baby girls given in Alberta in " + str(year) + \
          ":")
    print_girl_names(g_data) 
    
    print("\nTop 10 names for baby boys given in Alberta in " + str(year) \
          + ":")       
    print_boy_names(b_data)
    
def wildcard_search(names):
    '''
    Expects the names dictionary as a parameter. This function allows the user 
    to enter a name with an asterisk (*) representing missing letter(s). 
    This function is case insensitive. There are three parts to this function:
    a) names ending with an asterisk, e.g. franc*
    b) names starting with an asterisk, e.g. *elly
    c) names with an asterisk not at the beginning or the end, e.g. moh*had
    
    This function outsource the part of printing names to other functions,
    if name end with an asterisk: name_end_ast(full_list, search)
    if names starting with an asterisk: name_st_ast(full_list, search)
    if names with an asterisk not at the beginning or the end: 
    mid_ast(full_list, search)
    
    Parameters:  names: dictionary
                          dictionary of all the names
                          
    Return: NONE 
    '''
    
    if len(names.keys()) == 0:
        print("There are no data")
        return
    
    search = input("Enter name with * indicating missing letters: ")
    search = search.capitalize()
    
    # Turns this data {'Michael': [[732, 'Boy', 1980], [705, 'Boy', 1981]]} into
    # the following: {'Michael': {1980: [['Boy', 732]], 1981: [['Boy', 705]]
    full_list = {}                  # New dictionary
    for keys in names:              # keys refer to name in dictionary.
        full_list[keys] = {}        # add [key]-name dictionary to full-list 
        for lis in names[keys]:     # lis is the list data for single name.
            year = lis[-1]
            # if year has not been added, then add.
            if year not in full_list[keys].keys():
                full_list[keys][year] = []
            else: pass
            gender = lis[1]
            freq = lis[0]
            # Ex: {'Michael': {1980: [['Boy', 732], ['Girl', 705]]
            full_list[keys][year].append([gender, freq])
            
    # Passing the data to search helper for printing out the names.
    if "*" in search[-1]:
        name_end_ast(full_list, search)
        
    if "*" in search[0]:
        name_st_ast(full_list, search)
        
    elif not "*" in search[-1]:
        if "*" in search:
            mid_ast(full_list, search)
    
def name_end_ast(full_list, search):
    '''
    Expects the dictionary of all names in which data sorted is with separate 
    years, each year containning the gender and frequency and the name to be 
    searched as parameters. This function takes the names ending with an
    asterisk, e.g. franc*. Searches for that name is dictionary, and pass on 
    the results to all_print_helper(yr_dict) to print out the names.
    
    Parameters:
        full_list: dictionary
               dictionary of all names, data sorted with separate years.
        search: str
              name to be searched
    
    Return: NONE
    '''    
    s = search
    
    search = search.replace("*", "")
    search = search.capitalize()
    ser_len = len(search)
    match_names = []
    
    # for name in the full_list, that has the following data (as ex.)
    # {'Michael': {1980: [['Boy', 732], ['Girl', 705]]
    for list_name in full_list:
        if list_name[:ser_len] == search: # If name is same as being searched
            match_names.append(list_name)
            
    if len(match_names) == 0: print("No name found using " + s); return
    
    for n in match_names:
        yr_dict = full_list[n]
        yr_dict = dict(sorted(yr_dict.items()))    
        
        print("\n\tBoys\tGirls")        
        print(n)
        
        all_print_helper(yr_dict)
        
def all_print_helper(yr_dict):

    '''
    Expects dictionary with years as keys each containing list of list, with
    each list having gender and frequency.
    Ex: {1980: [['Boy', 256], ['Girl', 281]]}
        {1980: [['Boy', 732]]}
    This function takes the data, iterates to see type of it (if the name is 
    given to only boy or only girl or both) and finally prints the data 
    accordingly.
    Ex: 
    Michael:
                Boys	Girls
        1980 	732	 0
        
    This function is used in all three cases of wildcard search.
    
    Parameters:
        yr_dict: dictionary
               dictionary of all years, each conatining gender and their 
               respective frequency.
    
    Return: NONE
    '''
    # Acessing years in the keys.
    for years in (yr_dict.keys()):  
        # Try/except used to prevent 'index out of range error'
        try:
            # Case where name is given to only boys.
            # If Boy in {'Michael':{1980: [['Boy', 732], ['Girl', 705]]
            if  'Boy' in yr_dict[years][0]: 
                # if length of {1980: [['Boy', 732], ['Girl', 705]]} =1
                if len(yr_dict[years]) == 1:
                    print(str(years) + ":" +  "\t", end = "")
                    print(yr_dict[years][0][1], end = "")           
                    print("\t" + "0")
                            
            # Case where name is only given to girl
            # Ex: full_list[search][years][0] = ['Boy', 732]
            if 'Girl' in yr_dict[years][0]:
                if len(yr_dict[years]) == 1:
                    print(str(years) + ":" + "\t" + "0", end = "")
                    print("\t" + str(yr_dict[years][0][1]))    
                
            # Case where name is given to both boys and girls
            if 'Boy' in yr_dict[years][0] and \
            'Girl' in yr_dict[years][1]:
                print(str(years) + ":" +  "\t", end = "")
                print(yr_dict[years][0][1], end = "")
                print("\t" + str(yr_dict[years][1][1]))                     
        except: pass            

def name_st_ast(full_list, search):
    '''
    Expects the dictionary of all names in which data sorted is with separate 
    years, each year containning the gender and frequency and the name to be 
    searched as parameters. This function takes the names starting with an
    asterisk, e.g. *elly. Searches for that name is dictionary, and pass on 
    the results to all_print_helper(yr_dict) to print out the names.
    
    Parameters:
        full_list: dictionary
               dictionary of all names, data sorted with separate years.
        search: str
              name to be searched
    
    Return: NONE
    '''        
    s = search
    search = search.replace("*", "")
    ser_len = len(search) * (-1)
    match_names = []
    
    # for name in the full_list, that has the following data (as ex.)
    # {'Michael': {1980: [['Boy', 732], ['Girl', 705]]
    for list_name in full_list:
        if list_name[ser_len:] == search: #If the name is same as being searched
            match_names.append(list_name)
            
    if len(match_names) == 0: print("No name found using " + s); return
    
    for n in match_names:
        yr_dict = full_list[n]
        yr_dict = dict(sorted(yr_dict.items()))    
        
        print("\n\tBoys\tGirls")        
        print(n)
        
        all_print_helper(yr_dict)
        
def mid_ast(full_list, search):
    '''
    Expects the dictionary of all names in which data sorted is with separate 
    years, each year containning the gender and frequency and the name to be 
    searched as parameters. This function takes the name with asterisk not 
    at the beginning or the end. Ex: moh*had. Searches for that name is 
    dictionary, and pass on the results to all_print_helper(yr_dict) to print
    out the names.
    
    
    Parameters:
        full_list: dictionary
               dictionary of all names, data sorted with separate years.
        search: str
              name to be searched
    
    Return: NONE
    '''        
    s = search    
    sr_list = search.split("*")
    # string part before "*", string part after "*"
    st_str, end_str = sr_list[0], sr_list[1]
    st_str, end_str = st_str.capitalize(), end_str.lower()
    
    len_st_str, len_end_str = len(st_str), len(end_str) * (-1)
    match_names = []
    
    # for name in the full_list, that has the following data (as ex.)
    # {'Michael': {1980: [['Boy', 732], ['Girl', 705]]
    for list_name in full_list:
        if list_name[:len_st_str] == st_str and list_name[len_end_str:] \
           == end_str: #If the name is same as being searched
            match_names.append(list_name)
            
    if len(match_names) == 0: print("No name found using " + s); return
    
    for n in match_names:
        yr_dict = full_list[n]
        yr_dict = dict(sorted(yr_dict.items()))    
        
        print("\n\tBoys\tGirls")        
        print(n)
        
        all_print_helper(yr_dict)
       
def plot(names, year):
    '''
    Uses the matplotlib module to implement the trend graph.
    
     Parameters:
        names: dictionary
               dictionary of all names
        year: int
              latest year
              
    Return:  None
    '''    
    
    if len(names.keys()) == 0:
        print("There are no data")
        return
    
    search = input("Enter a name: ")
    search = search.capitalize()
    if search not in names.keys():
        print("There were no babies named " + search + " born in Alberta" 
              " between 1980 and " + str(year))
        return
    
    full_list = name_trend(names, year)    
    
    yr_dict = full_list[search]
    yr_dict = dict(sorted(yr_dict.items()))
    
    boys = {}
    girls = {}
    
    # for name in the full_list, that has the following data (as ex.)
    # {'Michael': {1980: [['Boy', 732], ['Girl', 705]]
    for list_name in full_list:
        if list_name == search:     # If the name is same as being searched
            # Acessing years in the keys.
            for years in (yr_dict.keys()):      
                try:
                    # Case where name is given to only boys.
                    # If Boy in {'Michael':{1980: [['Boy', 732], ['Girl', 705]]
                    if  'Boy' in yr_dict[years][0]: 
                        boys[years] = yr_dict[years][0][1]
                        girls[years] = 0
                            
                    # Case where name is only given to girl
                    # Ex: full_list[search][years][0] = ['Boy', 732]
                    if 'Girl' in yr_dict[years][0]:   
                        girls[years] = yr_dict[years][0][1]
                        boys[years] = 0
                        
                    # Case where name is given to both boys and girls
                    if 'Boy' in yr_dict[years][0] and \
                    'Girl' in yr_dict[years][1]:       
                        boys[years] = yr_dict[years][0][1]
                        girls[years] = yr_dict[years][1][1]                        
                       
                except: pass
    
    display_data(boys, girls, search , year)
    
def display_data(boys, girls, search , year):
    '''
    Uses the mathplotlob module to implement the trend graph.
    
     Parameters:
        boys: dictionary
               dictionary of the boys names with year as key and freq as value
        girls: dictionary
               dictionary of the girls names with year as key and freq as value
        search: str: name of the person 
        largest: int: largest freqency of boys/girl name
         year: int
              latest year
    Return:  None
    '''    
   
  
    
    # adding zero to when there's no name to prevent dimension error
    x_ticks = []
    years = []
    for yr in range(1980, year + 1):
        years.append(yr)
        
        if yr not in boys: boys[yr] = 0        
        if yr not in girls: girls[yr] = 0  
        
        yr = str(yr)        
        x_ticks.append(yr[2:])        
            
    boys_freq = list(boys.values())
    girls_freq = list(girls.values()) 
    

    boys_freq = np.array(boys_freq)
    girls_freq = np.array(girls_freq)
    
  
    my_xticks  = np.array(x_ticks)
    years = np.array(years)
    
    

    plt.ylabel('Frequency of Name')
    plt.xlabel('Years')
    plt.title(f"Trend for the name {search}")
    plt.xticks(years, my_xticks)
    plt.plot(years, girls_freq, label = "Girls")        
    plt.plot(years, boys_freq, label = "Boys")
    plt.legend()                
    plt.show()

    
def get_choice():
    '''
    This prompts user with "Select option (0 to 4): ", inputs, validates user 
    entry, and returns valid entry as an integer
    
    Parameters:  None
    Return:  choice: int: 
                   user selection
    '''
    
    choice = input("\nEnter command: ")
    # Error checking, number must be between 0 and 6.
    while True:
        if len(choice) != 0 and choice.isdigit():
            choice = int(choice)
            if 0 <= choice <= 7:
                return choice
            
        choice = input("Enter command: ")   
        
        
def main():
    """
    This displays a menu, prompts the user to enter a selection, and
    executes that selection.
    
    Parameters:  None
    Return    :  None
    """    
    names = {}
    top_ten = {}
    max_year = int()
    year = int()
    
    while True:
        print("\nAlberta Baby names\n"
              "-------------------------------\n"
              "(0) Quit\n"
              "(1) Load and process spreadsheet file\n"
              "(2) Save processed data\n"
              "(3) Open processed data\n"
              "(4) Search for a name\n"
              "(5) Print top ten list for a year\n"
              "(6) Search for names with specific letters\n"
              "(7) Graphically display the trend of a name\n")
        choice = get_choice()
    
        if   choice == 0: break
        elif choice == 1: 
            data = load_file("Baby_Names_Frequencies.xlsx", names, top_ten)
            if data != None:
                names, top_ten, max_year = data
        elif choice == 2: 
                data = (names, top_ten, max_year)
                pickle_helper(data)
        elif choice == 3: 
                data = load_helper()
                if data != None:
                    names, top_ten, max_year = data
                
        elif choice == 4: name_search(names, max_year)
        elif choice == 5: print_top_ten(top_ten, max_year)
        elif choice == 6: wildcard_search(names)
        elif choice == 7: plot(names, max_year)
        
    print("Goodbye")

if __name__ == "__main__":
    main()